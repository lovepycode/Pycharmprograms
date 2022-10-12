import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_90 = {}

#testing
 
 
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation for total value of inventory
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 90:
        products_under_90[product_no] = str(product_no) + " " + supplier_name + " " + str(inventory)
    # get all the products that have inventory under 90

    # this will update the excel sheet
    inventory_price.value = inventory * price
    inv_file.save("Inventory.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_90)
